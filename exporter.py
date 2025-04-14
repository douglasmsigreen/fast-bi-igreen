# exporter.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from typing import List, Tuple, Dict, Any # Adicionar Dict, Any
import logging
import io

logger = logging.getLogger(__name__)

class ExcelExporter:
    EXCEL_SETTINGS = {
        "header_color": "008001",
        "header_font_color": "FFFFFF",
        "freeze_panes": True,
    }
    CONDITIONAL_FORMATTING_TARGET_COLUMN = "Nome" # Coluna para formatação condicional

    # --- Métodos auxiliares (_add_headers, _add_data, _apply_formattings) ---
    def _add_headers(self, ws, headers: List[str]) -> None:
        """Adiciona cabeçalhos formatados a uma planilha."""
        if not headers: return
        ws.append(headers)
        fill = PatternFill(start_color=self.EXCEL_SETTINGS["header_color"], end_color=self.EXCEL_SETTINGS["header_color"], fill_type="solid")
        font = Font(bold=True, color=self.EXCEL_SETTINGS["header_font_color"])
        # Aplica estilo apenas às células da primeira linha (cabeçalho)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

    def _add_data(self, ws, data: List[tuple]) -> None:
        """Adiciona dados a uma planilha."""
        if not data: return
        logger.info(f"Adicionando {len(data)} linhas de dados à aba '{ws.title}'...")
        for i, row in enumerate(data):
            # Converte tudo para string para evitar problemas de tipo com openpyxl
            ws.append([str(v) if v is not None else "" for v in row])
        # logger.info(f"Dados adicionados à aba '{ws.title}'.") # Log menos verboso

    def _apply_formattings(self, ws, num_rows: int, headers: List[str]) -> None:
        """Aplica formatações (largura, filtro, congela painéis, condicional) a uma planilha."""
        logger.info(f"Aplicando formatações Excel na aba '{ws.title}'...")
        num_cols = len(headers)
        if not num_cols:
            logger.warning(f"Sem cabeçalhos para formatar na aba '{ws.title}'.")
            return

        # Congelar Painéis
        if self.EXCEL_SETTINGS["freeze_panes"]:
            ws.freeze_panes = ws['A2'] # Congela abaixo do cabeçalho

        # Auto Filtro
        if num_rows > 0: # Só aplica filtro se houver dados
             ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
        elif num_cols > 0: # Se não há dados, mas há colunas, aplica filtro só no header
             ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"

        # Ajustar Largura das Colunas (baseado em cabeçalho e amostra de dados)
        # logger.info(f"Ajustando largura das colunas para a aba '{ws.title}' (amostragem)...") # Log menos verboso
        for idx in range(1, num_cols + 1):
            letter = get_column_letter(idx)
            # Calcula o comprimento máximo inicial com base no cabeçalho
            header_len = len(str(headers[idx-1])) if headers[idx-1] else 0
            max_len = header_len
            # Verifica o comprimento nas primeiras 50 linhas de dados (ou menos se houver menos)
            for r_idx in range(2, min(num_rows + 2, 52)):
                try:
                    cell_val = ws.cell(row=r_idx, column=idx).value
                    # Considera o comprimento apenas se a célula tiver valor
                    if cell_val: max_len = max(max_len, len(str(cell_val)))
                except IndexError: pass # Ignora se a linha não existir
            # Define a largura ajustada com limites mínimo e máximo
            adjusted_width = min(max(max_len + 2, 10), 60) # Min 10, Max 60
            ws.column_dimensions[letter].width = adjusted_width

        # Formatação Condicional (se houver dados)
        if num_rows > 0:
            target = self.CONDITIONAL_FORMATTING_TARGET_COLUMN
            try:
                # Encontra o índice da coluna alvo (base 1)
                target_idx = headers.index(target) + 1
                target_letter = get_column_letter(target_idx)
                logger.info(f"Aplicando formatação condicional na coluna {target_letter} ('{target}') da aba '{ws.title}'...")
                # Define a regra de escala de cores
                rule = ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color=self.EXCEL_SETTINGS["header_color"])
                # Adiciona a regra ao intervalo de dados da coluna alvo
                ws.conditional_formatting.add(f"{target_letter}2:{target_letter}{num_rows + 1}", rule)
            except ValueError:
                # Loga um aviso se a coluna alvo não for encontrada nos cabeçalhos
                logger.warning(f"Cabeçalho '{target}' para formatação condicional não encontrado na aba '{ws.title}'.")
            except Exception as e:
                # Loga um aviso se ocorrer outro erro durante a formatação condicional
                logger.warning(f"Erro ao aplicar formatação condicional na aba '{ws.title}': {e}", exc_info=False)

        # logger.info(f"Formatações Excel aplicadas na aba '{ws.title}'.") # Log menos verboso

    # --- Método para Exportação de Planilha Única ---
    def export_to_excel_bytes(self, data: List[tuple], headers: List[str], sheet_name: str = "Clientes") -> bytes:
        """Gera UM ficheiro Excel com UMA aba e retorna como bytes."""
        if not headers:
            raise ValueError("Cabeçalhos não fornecidos para exportação de aba única.")
        try:
            logger.info(f"Iniciando geração de Excel (aba única: {sheet_name}) com {len(data)} registros...")
            wb = openpyxl.Workbook(write_only=False) # write_only=False para permitir formatação
            ws = wb.active
            ws.title = sheet_name

            self._add_headers(ws, headers)
            self._add_data(ws, data)
            self._apply_formattings(ws, len(data), headers)

            excel_bytes_io = io.BytesIO()
            wb.save(excel_bytes_io)
            excel_bytes_io.seek(0) # Volta ao início do stream de bytes
            logger.info(f"Ficheiro Excel (aba única: {sheet_name}) gerado em memória.")
            return excel_bytes_io.getvalue()

        except Exception as e:
            logger.error(f"Erro durante a geração do Excel (aba única) ({type(e).__name__}): {e}", exc_info=True)
            raise RuntimeError(f"Erro ao gerar o ficheiro Excel (aba única): {e}")

    # --- Método para Exportação com Múltiplas Abas ---
    def export_multi_sheet_excel_bytes(self, sheets_data: List[Dict[str, Any]]) -> bytes:
        """
        Gera UM ficheiro Excel com MÚLTIPLAS abas e retorna como bytes.
        sheets_data: Lista de dicionários, cada um com chaves 'name', 'headers', 'data'.
                     Ex: [{'name': 'Aba1', 'headers': [...], 'data': [...]}, ...]
        """
        if not sheets_data:
            raise ValueError("Nenhuma informação de aba fornecida para exportação multi-abas.")

        try:
            logger.info(f"Iniciando geração de Excel multi-abas ({len(sheets_data)} abas)...")
            wb = openpyxl.Workbook(write_only=False) # write_only=False para permitir formatação
            # Remove a aba padrão ("Sheet") criada automaticamente
            if "Sheet" in wb.sheetnames:
                 default_sheet = wb["Sheet"]
                 wb.remove(default_sheet)

            # Itera sobre os dados de cada aba a ser criada
            for sheet_info in sheets_data:
                sheet_name = sheet_info.get('name', f'Aba_{len(wb.sheetnames)+1}') # Nome padrão se não fornecido
                headers = sheet_info.get('headers', [])
                data = sheet_info.get('data', [])

                # Pula a criação da aba se não houver cabeçalhos definidos
                if not headers:
                     logger.warning(f"Pulando aba '{sheet_name}' por falta de cabeçalhos.")
                     continue

                logger.info(f"Criando aba: '{sheet_name}' com {len(data)} registros.")
                # Cria a nova aba no workbook
                ws = wb.create_sheet(title=sheet_name)

                # Adiciona cabeçalhos, dados e aplica formatações usando os métodos auxiliares
                self._add_headers(ws, headers)
                self._add_data(ws, data) # Lida com dados vazios internamente
                self._apply_formattings(ws, len(data), headers)

            # Verifica se alguma aba foi realmente criada
            if not wb.sheetnames:
                 logger.warning("Nenhuma aba foi criada no Excel multi-abas (talvez todas sem cabeçalhos).")
                 # Cria uma aba vazia para evitar erro ao salvar um workbook sem abas
                 wb.create_sheet(title="Vazio")

            # Salva o workbook completo em um stream de bytes na memória
            excel_bytes_io = io.BytesIO()
            wb.save(excel_bytes_io)
            excel_bytes_io.seek(0) # Volta ao início do stream
            logger.info("Ficheiro Excel multi-abas gerado em memória.")
            return excel_bytes_io.getvalue()

        except Exception as e:
            logger.error(f"Erro durante a geração do Excel (multi-abas) ({type(e).__name__}): {e}", exc_info=True)
            # Propaga o erro para ser tratado na rota Flask
            raise RuntimeError(f"Erro ao gerar o ficheiro Excel (multi-abas): {e}")