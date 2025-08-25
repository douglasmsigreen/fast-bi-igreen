# backend/exporter.py
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Classe utilitária para gerar arquivos Excel a partir de dados de consultas.
    """

    def __init__(self):
        self.wb = Workbook()

    def _add_headers(self, ws, headers: List[str]):
        """Adiciona os cabeçalhos em uma planilha e aplica formatação."""
        header_fill = PatternFill(start_color="3C8DBC", end_color="3C8DBC", fill_type="solid")
        header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}1"]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 20
        return ws

    def _add_data(self, ws, data: List[List[Any]]):
        """Adiciona os dados em uma planilha, ajusta a largura das colunas e formata."""
        if not data:
            return ws

        # Supondo que 'data' é uma lista de listas de valores
        for row_data in data:
            ws.append(row_data)

        # Resto da formatação
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
        
        # Ajuste da largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column letter
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError):
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width if adjusted_width < 50 else 50
        return ws

    def export_to_excel_bytes(self, data: List[List[Any]], headers: List[str], sheet_name: str = "Sheet1") -> bytes:
        """
        Gera um arquivo Excel de aba única em memória (bytes).
        Args:
            data (list of lists): Os dados a serem inseridos.
            headers (list): A lista de cabeçalhos.
            sheet_name (str): O nome da aba.
        Returns:
            bytes: O conteúdo do arquivo Excel em formato de bytes.
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            self.wb = wb
            self._add_headers(ws, headers)
            self._add_data(ws, data)
            
            buffer = BytesIO()
            self.wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        except Exception as e:
            logger.error(f"Erro ao gerar o ficheiro Excel (aba única): {e}", exc_info=True)
            raise RuntimeError(f"Erro ao gerar o ficheiro Excel (aba única): {e}")

    def export_multi_sheet_excel_bytes(self, sheets: List[Dict[str, Any]]) -> bytes:
        """
        Gera um arquivo Excel com múltiplas abas em memória (bytes).
        Args:
            sheets (list of dict): Uma lista de dicionários, onde cada um representa uma aba
                                    com 'name', 'headers' e 'data'.
        Returns:
            bytes: O conteúdo do arquivo Excel em formato de bytes.
        """
        try:
            wb = Workbook()
            default_ws = wb.active
            wb.remove(default_ws) # Remove a aba padrão para começar limpo
            
            for sheet_info in sheets:
                ws = wb.create_sheet(title=sheet_info['name'])
                self._add_headers(ws, sheet_info['headers'])
                self._add_data(ws, sheet_info['data'])
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        except Exception as e:
            logger.error(f"Erro ao gerar o ficheiro Excel (multi-aba): {e}", exc_info=True)
            raise RuntimeError(f"Erro ao gerar o ficheiro Excel (multi-aba): {e}")